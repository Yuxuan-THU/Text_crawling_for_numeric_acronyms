import openpyxl

# 读取 Excel 文件
wb = openpyxl.load_workbook("ex.xlsx")  # 根据实际表格文件名进行调整
ws = wb.active

# 遍历表格的第一列
rows_to_delete = []
for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
    word = str(row[0].value)  # 获取第一列的单元格的值并转换为字符串
    if not word[0].isdigit():  # 判断第一个字符是否是汉字数字
        rows_to_delete.append(row[0].row)  # 如果不是，则记录需要删除的行号

# 倒序删除需要删除的行，避免删除行导致索引变化问题
for row in reversed(rows_to_delete):
    ws.delete_rows(row)

# 保存处理后的表格
wb.save("result.xlsx")  # 根据实际需求选择合适的保存文件名和文件类型

